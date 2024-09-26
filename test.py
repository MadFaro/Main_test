import os
import shutil
from openpyxl import load_workbook, Workbook

def move_folders_from_excel(excel_file, archive_path, output_file):
    # Открываем Excel-файл с путями
    wb = load_workbook(excel_file)
    ws = wb.active

    # Проверяем, что папка для архива существует, если нет, создаем её
    if not os.path.exists(archive_path):
        os.makedirs(archive_path)

    # Создаем новый Excel-файл для записи результатов
    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = "Move Report"

    # Добавляем заголовки в новый файл
    result_ws.append(["Folder Path", "Status", "Error (if any)"])

    # Проходим по строкам Excel-файла, начиная со второй строки (первая — заголовки)
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        folder_path = row[0]  # Путь к папке находится в первом столбце

        if folder_path and os.path.exists(folder_path):
            # Получаем имя папки для перемещения
            folder_name = os.path.basename(folder_path)

            # Путь к новой папке в архиве
            destination = os.path.join(archive_path, folder_name)

            try:
                # Перемещаем папку
                shutil.move(folder_path, destination)
                status = f"Папка перемещена в {destination}"
                error = ""
            except Exception as e:
                status = "Ошибка при перемещении"
                error = str(e)

            # Записываем результат в новый Excel-файл
            result_ws.append([folder_path, status, error])

        else:
            status = "Папка не найдена"
            error = ""
            result_ws.append([folder_path, status, error])

    # Сохраняем результаты в новый Excel-файл
    result_wb.save(output_file)

# Пример использования
if __name__ == "__main__":
    excel_file = "! Разбор папок.xlsx"  # Укажите имя вашего Excel файла
    archive_path = r"V:\VOL2\Contact-center\Файлы\Аналитика\Архив"  # Путь к папке, куда нужно переместить папки
    output_file = "move_report.xlsx"  # Имя файла для записи результатов

    move_folders_from_excel(excel_file, archive_path, output_file)

    print(f"Результаты сохранены в файл: {output_file}")

